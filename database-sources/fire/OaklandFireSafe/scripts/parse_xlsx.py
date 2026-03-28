"""
Parse the Oakland Fire Safe Council (OFSC) fire-resistant/fire-prone plant lists.

Source: City of Oakland Public Works Watershed Division, FireSafe Marin,
        and Diablo FireSafe Council.
File: Fire-Resistant-Fire-Prone-Lists_OFSC.xlsx

Three sheets:
  1. Fire Resistant (239 rows) - plants with fire resistance
  2. Fire Prone (64 rows) - plants that are fire hazards
  3. Local Botan Gardns & Natives - reference links (not extracted)
"""

import csv
import json
import os
import sqlite3

from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.dirname(SCRIPT_DIR)
XLSX_FILE = os.path.join(DATA_DIR, "Sources",
                         "Fire-Resistant-Fire-Prone-Lists_OFSC.xlsx")


def parse_fire_resistant(wb):
    """Parse the Fire Resistant sheet."""
    ws = wb["Fire Resistant "]
    plants = []
    current_category = ""

    for row in ws.iter_rows(min_row=4, values_only=True):  # skip header rows
        vals = list(row)
        sci_name = vals[0]
        common_name = vals[1]

        if not sci_name and not common_name:
            continue

        # Category headers (e.g., "GROUND COVERS", "SHRUBS", "TREES")
        if sci_name and not common_name and sci_name == sci_name.upper():
            current_category = sci_name.strip()
            continue

        if not common_name:
            continue

        ca_native = str(vals[4]).strip() if vals[4] else ""
        is_native = ca_native == "X"
        origin = "California Native" if is_native else (ca_native if ca_native else "")

        plants.append({
            "scientific_name": str(sci_name).strip() if sci_name else "",
            "common_name": str(common_name).strip(),
            "lifeform": str(vals[2]).strip() if vals[2] else "",
            "category": current_category,
            "comments": str(vals[3]).strip().replace("\n", " ") if vals[3] else "",
            "ca_native": is_native,
            "origin": origin,
            "fire_rating": "Fire-Resistant",
            "recommendation": "",
            "pyrophytic": False,
        })

    return plants


def parse_fire_prone(wb):
    """Parse the Fire Prone sheet."""
    ws = wb["Fire Prone"]
    plants = []
    current_category = ""

    for row in ws.iter_rows(min_row=5, values_only=True):  # skip header rows
        vals = list(row)
        pyro = str(vals[0]).strip() if vals[0] else ""
        sci_name = vals[1]
        common_name = vals[2]

        if not sci_name and not common_name:
            continue

        # Category headers
        if sci_name and not common_name and str(sci_name) == str(sci_name).upper():
            current_category = str(sci_name).strip()
            continue

        if not common_name:
            continue

        ca_native = str(vals[5]).strip() if vals[5] else ""
        is_native = ca_native == "X"

        plants.append({
            "scientific_name": str(sci_name).strip() if sci_name else "",
            "common_name": str(common_name).strip(),
            "lifeform": str(vals[3]).strip() if vals[3] else "",
            "category": current_category,
            "comments": "",
            "ca_native": is_native,
            "origin": "California Native" if is_native else "",
            "fire_rating": "Fire-Prone",
            "recommendation": str(vals[4]).strip() if vals[4] else "",
            "pyrophytic": pyro == "F",
        })

    return plants


def main():
    print(f"Reading: {XLSX_FILE}")
    wb = load_workbook(XLSX_FILE, read_only=True)

    resistant = parse_fire_resistant(wb)
    print(f"Fire-Resistant plants: {len(resistant)}")

    prone = parse_fire_prone(wb)
    print(f"Fire-Prone plants: {len(prone)}")

    wb.close()

    all_plants = resistant + prone
    print(f"Total plants: {len(all_plants)}")

    # --- CSV ---
    csv_path = os.path.join(DATA_DIR, "plants.csv")
    fields = ["scientific_name", "common_name", "lifeform", "category",
              "comments", "ca_native", "origin", "fire_rating",
              "recommendation", "pyrophytic"]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for p in all_plants:
            writer.writerow(p)
    print(f"Wrote {csv_path}")

    # --- JSON ---
    json_path = os.path.join(DATA_DIR, "plants.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_plants, f, indent=2, ensure_ascii=False)
    print(f"Wrote {json_path}")

    # --- SQLite ---
    db_path = os.path.join(DATA_DIR, "plants.db")
    if os.path.exists(db_path):
        os.remove(db_path)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE plants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            scientific_name TEXT,
            common_name TEXT,
            lifeform TEXT,
            category TEXT,
            comments TEXT,
            ca_native BOOLEAN,
            origin TEXT,
            fire_rating TEXT,
            recommendation TEXT,
            pyrophytic BOOLEAN
        )
    """)
    cur.execute("CREATE INDEX idx_scientific_name ON plants(scientific_name)")
    cur.execute("CREATE INDEX idx_fire_rating ON plants(fire_rating)")
    cur.execute("CREATE INDEX idx_ca_native ON plants(ca_native)")

    for p in all_plants:
        cur.execute(
            """INSERT INTO plants (scientific_name, common_name, lifeform, category,
               comments, ca_native, origin, fire_rating, recommendation, pyrophytic)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (p["scientific_name"], p["common_name"], p["lifeform"], p["category"],
             p["comments"], p["ca_native"], p["origin"], p["fire_rating"],
             p["recommendation"], p["pyrophytic"]),
        )
    conn.commit()

    # Stats
    cur.execute("SELECT fire_rating, COUNT(*) FROM plants GROUP BY fire_rating")
    print("\nFire Rating Distribution:")
    for rating, count in cur.fetchall():
        print(f"  {rating}: {count}")

    cur.execute("SELECT category, fire_rating, COUNT(*) FROM plants GROUP BY category, fire_rating ORDER BY fire_rating, COUNT(*) DESC")
    print("\nCategory Breakdown:")
    for cat, rating, count in cur.fetchall():
        print(f"  {rating} / {cat}: {count}")

    cur.execute("SELECT COUNT(*) FROM plants WHERE ca_native = 1")
    print(f"\nCA Native: {cur.fetchone()[0]}")

    cur.execute("SELECT COUNT(*) FROM plants WHERE pyrophytic = 1")
    print(f"Pyrophytic (extremely flammable): {cur.fetchone()[0]}")

    cur.execute("SELECT recommendation, COUNT(*) FROM plants WHERE recommendation != '' GROUP BY recommendation")
    print("\nFire-Prone Recommendations:")
    for rec, count in cur.fetchall():
        print(f"  {rec}: {count}")

    conn.close()
    print(f"\nWrote {db_path}")


if __name__ == "__main__":
    main()
